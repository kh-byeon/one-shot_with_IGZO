
from silence_tensorflow import silence_tensorflow
silence_tensorflow()

import tracemalloc
import tensorflow as tf
#from tensorflow.python.framework.ops import disable_eager_execution
#disable_eager_execution()

import os,shutil
import csv
from openpyxl import load_workbook, Workbook
os.environ["CUDA_VISIBLE_DEVICES"]="0"
#from IGZO_TCAM_training import SiameseNetworkforTCAM
#from pre_trained_modified_siamese_network import SiameseNetworkforTCAM
#from vgg_19 import SiameseNetworkforTCAM
#from modified_siamese_network import SiameseNetworkforTCAM
from current_training import SiameseNetworkforCurrent_training

#from encoder_network import EncdoerNetwork
import numpy as np
np.set_printoptions(precision=3, suppress=True)
np.set_printoptions(threshold=np.inf, linewidth=np.inf)
gpus = tf.config.experimental.list_physical_devices('GPU')


if gpus:
  try:
    tf.config.experimental.set_memory_growth(gpus[0], True)
  except RuntimeError as e:    
    print(e)

def main(lr):
    dataset_path = '../omniglot/python'
    use_augmentation = True
    learning_rate = lr
    batch_size = 32
    margin= 3
      

    # Learning Rate multipliers for each layer
    learning_rate_multipliers = {}
    learning_rate_multipliers['Conv1'] = 1
    learning_rate_multipliers['Conv2'] = 1
    learning_rate_multipliers['Conv3'] = 1
    learning_rate_multipliers['Conv4'] = 1
    learning_rate_multipliers['Conv5'] = 1
    learning_rate_multipliers['Conv6'] = 1
    learning_rate_multipliers['Dense1'] = 1
    learning_rate_multipliers['Dense2'] = 1

    #l2-regularization penalization for each layer
    l2_penalization = {}
    l2_penalization['Conv1'] = 0
    l2_penalization['Conv2'] = 0
    l2_penalization['Conv3'] = 0
    l2_penalization['Conv4'] = 0
    l2_penalization['Conv5'] = 0
    l2_penalization['Conv6'] = 0
    l2_penalization['Dense1'] =0
    l2_penalization['Dense2'] = 0

    # Path where the logs will be saved
    tensorboard_log_path = './logs/siamese_net'
    
    #siamese_network = VGG_19(
    siamese_network = SiameseNetworkforCurrent_training(    
    #siamese_network = SiameseNetworkforTCAM(    
    #siamese_network = EncoderNetwork(
        dataset_path=dataset_path,
        learning_rate=learning_rate,
        batch_size=batch_size, use_augmentation=use_augmentation,
        learning_rate_multipliers=learning_rate_multipliers,
        l2_regularization_penalization=l2_penalization,
        tensorboard_log_path=tensorboard_log_path,
        margin=margin
    )

    # Final layer-wise momentum (mu_j in the paper)
    momentum = 0.9
    # linear epoch slope evolution
    momentum_slope = 0.01
    support_set_size = 5
    validate_each = 1000
    evaluate_each = 2000
    number_of_train_iterations = 1000000
    #siamese_network.model_train.load_weights('./one_sigma_model/one_sigma.h5')
    #siamese_network.model_train.load_weights('./fitted_model/01per1st/siamese_net.h5')
    if not os.path.exists('./models'):
        os.makedirs('./models')
    if not os.path.exists('./logs'):
        os.makedirs('./logs')
    #shutil.rmtree('./models/')
    #shutil.rmtree('./logs/')
    if not os.path.exists('./bestmodel'):
        os.makedirs('./bestmodel')

    """
    validation_accuracy,acc_list = siamese_network.train_siamese_network(number_of_iterations=number_of_train_iterations,
                                                                support_set_size=support_set_size,
                                                                final_momentum=momentum,
                                                                momentum_slope=momentum_slope,
                                                                validate_each=validate_each, 
                                                                evaluate_each=evaluate_each,
                                                                model_name='one_sigma')
    
    
    if validation_accuracy == 0:
        evaluation_accuracy = 0
    else:
        # Load the weights with best validation accuracy
        siamese_network.model_test.load_weights('./models/one_sigma.h5')
        #evaluation_accuracy = siamese_network.omniglot_loader.one_shot_test(siamese_network.model_test,
        #evaluation_accuracy = siamese_network.omniglot_loader.one_shot_test_TCAM(siamese_network.model_test,
        
        evaluation_accuracy = siamese_network.omniglot_loader.one_shot_test_fitted_simmem(siamese_network.model_test,
                                                                        support_set_size, 40, False)
    
    print('Final Evaluation Accuracy = ' + str(evaluation_accuracy))
    """
    siamese_network.omniglot_loader.split_train_datasets()
    siamese_network.model_test.load_weights('./one_sigma_model/one_sigma.h5')
    evaluation_accuracy = siamese_network.omniglot_loader.one_shot_test_fitted_simmem(siamese_network.model_test,
                                                                        support_set_size, 40, False)
    if not os.path.exists('./lrsweep'):
        os.makedirs('./lrsweep')
     


    # File name
    file_name = "pre_trained_six_sigma_eval.xlsx"

    
    # Check if the file exists and load it, otherwise create a new one
    try:
        workbook = load_workbook(file_name)  
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook() 
        sheet = workbook.active
        sheet.title = "Evaluation Results"
        sheet.append(["Evaluation Accuracy"])  

    # Find the next empty row
    next_row = sheet.max_row + 1

    # Write the value to the next available row
    sheet.cell(row=next_row, column=1, value=evaluation_accuracy)

    # Save the workbook
    workbook.save(file_name)
    print(f"Appended {evaluation_accuracy} to {file_name}")


    
if __name__ == "__main__":
    #lr_list=[100,1,1e-4,1e-9]
    #lr_list=[100,50,10,1,1e-2,1e-4,1e-6,1e-9]
    #lr_list=[1e-4,1e-6,1e-9]
    #lr_list=[1e-9]
    lr_list=[1,1,1,1,1,1,1,1,1,1]
   
    
    if not os.path.exists('./lrsweep'):
        os.makedirs('./lrsweep')
    shutil.rmtree('./lrsweep/')
    
    for sweep in lr_list:
        main(sweep)


    print('type tensorboard --logdir=./logs/   for tensorboard')
    